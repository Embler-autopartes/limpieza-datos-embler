#!/usr/bin/env python3
"""
Genera los archivos para Matrixify del rollout TecDoc:
  1) collections_sistema.xlsx  -> nuevas Smart Collections (marca x sistema), regla _brand + tecdoc.sistema
  2) menu_pase1.xlsx           -> mega menu Marca -> Sistema -> Subgrupo (titles prefijados, REPLACE)
                                  Sistema linkea a la collection nueva; Subgrupo a la collection existente.
Estructura LEAN: top 8 sistemas/marca + top 3 subgrupos/sistema (por volumen).
NO toca Shopify; solo crea archivos.
"""
import csv, json, re, unicodedata
from collections import defaultdict
from pathlib import Path
import openpyxl

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
SRC_COLLS = ROOT / "Procesamiento de Catálogo/outputs/collections-matrixify/source/Embler-Collections.xlsx"
ARBOL = json.load(open(ROOT / "Procesamiento de Catálogo/outputs/_arbol_marca_grupo_subgrupo.json", encoding="utf-8"))

CAP_SIS, CAP_SUB = 8, 3
MENU_TITLE, MENU_HANDLE = "Mega menu", "mega-menu"
# (display, brand_metafield_value, brand_handle)
BRANDS = [("BMW","BMW","bmw"),("Mercedes Benz","Mercedes-Benz","mercedes-benz"),("Audi","Audi","audi"),
          ("Mini Cooper","Mini","mini"),("Porsche","Porsche","porsche"),("Smart","Smart","smart"),
          ("Volkswagen","Volkswagen","volkswagen"),("Bentley","Bentley","bentley"),("Fiat","Fiat","fiat"),
          ("Jaguar","Jaguar","jaguar"),("Land Rover","Land Rover","land-rover"),("Seat","Seat","seat"),
          ("Volvo","Volvo","volvo")]

def slug(s):
    s = unicodedata.normalize("NFKD", s).encode("ascii","ignore").decode().lower()
    return re.sub(r"-+","-", re.sub(r"[^a-z0-9]+","-", s)).strip("-")

# (grupo,subgrupo) -> sistema
SIS = {}
for r in csv.DictReader(open(HERE/"homologacion.csv", encoding="utf-8")):
    SIS[(r["grupo_embler"], r["subgrupo_embler"])] = r["tecdoc_sistema"]

# leaf handle lookup desde collections existentes: (brand,group,subgroup)->handle
wb = openpyxl.load_workbook(SRC_COLLS, data_only=True)
ws = wb["Smart Collections"]
coll = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    h, rc, _, rv = row[0], row[11], row[12], row[13]
    if not h: continue
    c = coll.setdefault(h, {"b":None,"g":None,"s":None})
    if rc=="Metafield: global._brand": c["b"]=rv
    elif rc=="Metafield: global.group": c["g"]=rv
    elif rc=="Metafield: global.sub_group": c["s"]=rv
leaf_handle = {(c["b"],c["g"],c["s"]): h for h,c in coll.items() if c["b"] and c["g"] and c["s"]}

# construir menu lean por marca: sistema -> [(grupo,subgrupo,count,leaf_handle)]
menu = {}  # brand_value -> ordered list of (sistema, [(sub_label, leaf_handle), ...])
sistema_set = defaultdict(set)  # brand_value -> {sistema}
for disp, bval, bh in BRANDS:
    gs = ARBOL.get(bval)
    if not gs: continue
    sis = defaultdict(list)  # sistema -> [(count, grupo, sub)]
    sis_tot = defaultdict(int)
    for g, subs in gs.items():
        for sg, n in subs.items():
            s = SIS.get((g, sg))
            if not s: continue
            sis[s].append((n, g, sg)); sis_tot[s]+=n
    top_sis = sorted(sis_tot, key=lambda x:-sis_tot[x])[:CAP_SIS]
    rows = []
    for s in top_sis:
        leaves = sorted(sis[s], key=lambda x:-x[0])[:CAP_SUB]
        leaf_list = []
        for n, g, sg in leaves:
            lh = leaf_handle.get((bval, g, sg))
            if lh: leaf_list.append((sg, lh))
        rows.append((s, leaf_list))
        sistema_set[bval].add(s)
    menu[bval] = rows

# ---------- 1) collections_sistema.xlsx ----------
COLL_COLS = ["Handle","Command","Title","Published","Must Match",
             "Rule: Product Column","Rule: Relation","Rule: Condition"]
wb1 = openpyxl.Workbook(); s1 = wb1.active; s1.title = "Smart Collections"; s1.append(COLL_COLS)
sis_coll_handle = {}  # (bval,sistema)->handle
created = 0
for disp, bval, bh in BRANDS:
    for s in sorted(sistema_set.get(bval, [])):
        ch = f"{bh}-sis-{slug(s)}"
        sis_coll_handle[(bval,s)] = ch
        title = f"{disp} {s}"
        # fila 1: brand rule + metadata
        s1.append([ch,"MERGE",title,"TRUE","all","Metafield: global._brand","Equals",bval])
        # fila 2: sistema rule
        s1.append([ch,"","","","","Metafield: tecdoc.sistema","Equals",s])
        created += 1
wb1.save(HERE/"collections_sistema.xlsx")

# ---------- 2) menu_pase1.xlsx (titles prefijados, REPLACE) ----------
MENU_COLS = ["Handle","Command","Title","Menu Item: Title","Menu Item: Command",
             "Menu Item: Resource Type","Menu Item: Resource Handle","Menu Item: URL",
             "Menu Item: Parent Title","Menu Item: Position"]
wb2 = openpyxl.Workbook(); s2 = wb2.active; s2.title = "Menus"; s2.append(MENU_COLS)
def mrow(it, rtype, rhandle, url, parent, pos):
    s2.append([MENU_HANDLE,"REPLACE",MENU_TITLE,it,"MERGE",rtype,rhandle,url,parent,pos])
items = 0
for pb,(disp,bval,bh) in enumerate(BRANDS, start=1):
    rows = menu.get(bval)
    if not rows: continue
    mrow(disp,"HTTP","","#","",pb); items+=1               # N1 marca (sin link)
    for ps,(s,leaves) in enumerate(rows, start=1):
        sis_title = f"{disp} - {s}"
        mrow(sis_title,"COLLECTION",sis_coll_handle[(bval,s)],"",disp,ps); items+=1  # N2 sistema (link a coll nueva)
        for pl,(sg,lh) in enumerate(leaves, start=1):
            mrow(f"{disp} - {s} - {sg}","COLLECTION",lh,"",sis_title,pl); items+=1   # N3 subgrupo (link existente)
wb2.save(HERE/"menu_pase1.xlsx")

print(f"collections_sistema.xlsx: {created} collections nuevas (marca x sistema)")
print(f"menu_pase1.xlsx: {items} items de menu")
print("Listos en base-pruebas/. Pase 2 (rename a labels cortos) se genera tras importar el pase 1.")
