"""
Procesa el resultado de import de Matrixify (Import_Result_2026-05-13_074131.xlsx)
y genera tres archivos:

1. outputs/2026-05-13-collections-errores/Embler-Collections-Errores.xlsx
   - Solo las 7 handles que fallaron (collection ya existe en Shopify).
   - Incluye comentario original de Matrixify para documentar.

2. outputs/2026-05-13-collections-bloques/Embler-Collections-Bloque-2.xlsx
   - Siguientes 300 handles del source (filas 301-600).

3. outputs/2026-05-13-collections-bloques/Embler-Collections-Bloque-3.xlsx
   - Handles restantes (filas 601-895).

Los bloques mantienen exactamente la misma estructura que el archivo source
(Embler-Collections.xlsx) para que Matrixify los acepte sin cambios.
"""

import openpyxl
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
SOURCE = ROOT / "outputs/2026-05-12-collections-matrixify/Embler-Collections.xlsx"
RESULT = ROOT / "Import_Result_2026-05-13_074131.xlsx"
OUT_ERRORES = ROOT / "outputs/2026-05-13-collections-errores/Embler-Collections-Errores.xlsx"
OUT_BLOQUE2 = ROOT / "outputs/2026-05-13-collections-bloques/Embler-Collections-Bloque-2.xlsx"
OUT_BLOQUE3 = ROOT / "outputs/2026-05-13-collections-bloques/Embler-Collections-Bloque-3.xlsx"

SHEET = "Smart Collections"


def load_source_rows():
    """Carga el source con todas las filas (handle -> lista de filas)."""
    wb = openpyxl.load_workbook(SOURCE, data_only=False)
    ws = wb[SHEET]
    headers = [c.value for c in ws[1]]
    rows_by_handle = defaultdict(list)
    handle_order = []
    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        h = row[0]
        if not h:
            continue
        if h not in seen:
            seen.add(h)
            handle_order.append(h)
        rows_by_handle[h].append(row)
    return headers, handle_order, rows_by_handle


def load_failed_handles():
    """Lee Import_Result y devuelve {handle: (title, comment)} para los Failed."""
    wb = openpyxl.load_workbook(RESULT, data_only=True)
    ws = wb[SHEET]
    failed = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        h, _cmd, title = row[0], row[1], row[2]
        result, comment = row[16], row[17]
        if result == "Failed" and h and h not in failed:
            failed[h] = (title, comment)
    return failed


def write_xlsx(path, headers, rows, extra_headers=None, extra_rows=None):
    """Escribe un xlsx con los headers + filas. extra_* se anexan al final."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET
    all_headers = headers + (extra_headers or [])
    ws.append(all_headers)
    for row in rows:
        ws.append(row)
    if extra_rows:
        for row in extra_rows:
            ws.append(row)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main():
    headers, handle_order, rows_by_handle = load_source_rows()
    failed = load_failed_handles()

    print(f"Source: {len(handle_order)} handles únicos.")
    print(f"Failed handles: {len(failed)}")

    # 1) Excel de errores: solo los 7 failed con info extra de Matrixify
    errores_rows = []
    for h, (title, comment) in failed.items():
        # Tomamos todas las filas del handle (puede tener varias reglas)
        for src_row in rows_by_handle[h]:
            errores_rows.append(list(src_row) + [comment, title])
    write_xlsx(
        OUT_ERRORES,
        headers,
        errores_rows,
        extra_headers=["Matrixify Error", "Matrixify Title"],
    )
    print(f"Errores -> {OUT_ERRORES.relative_to(ROOT)} ({len(errores_rows)} filas, {len(failed)} handles)")

    # Determinar handles pendientes (los que NO se intentaron)
    # = todos los del source que no aparecen en Import_Result
    wb_result = openpyxl.load_workbook(RESULT, data_only=True)
    ws_result = wb_result[SHEET]
    attempted = set()
    for row in ws_result.iter_rows(min_row=2, values_only=True):
        h = row[0]
        if h and isinstance(h, str) and not h.startswith("###") and h.strip():
            attempted.add(h)

    pending = [h for h in handle_order if h not in attempted]
    print(f"Pendientes: {len(pending)}")

    # 2) Bloque 2: primeros 300 pendientes
    bloque2_handles = pending[:300]
    bloque2_rows = []
    for h in bloque2_handles:
        bloque2_rows.extend(list(r) for r in rows_by_handle[h])
    write_xlsx(OUT_BLOQUE2, headers, bloque2_rows)
    print(f"Bloque 2 -> {OUT_BLOQUE2.relative_to(ROOT)} ({len(bloque2_handles)} handles, {len(bloque2_rows)} filas)")

    # 3) Bloque 3: resto
    bloque3_handles = pending[300:]
    bloque3_rows = []
    for h in bloque3_handles:
        bloque3_rows.extend(list(r) for r in rows_by_handle[h])
    write_xlsx(OUT_BLOQUE3, headers, bloque3_rows)
    print(f"Bloque 3 -> {OUT_BLOQUE3.relative_to(ROOT)} ({len(bloque3_handles)} handles, {len(bloque3_rows)} filas)")


if __name__ == "__main__":
    main()
