"""
Segundo paso del mega menu: renombra los items a labels CORTOS sin tocar la
jerarquía.

Como funciona:
1. Lee un export reciente del Mega menu de Matrixify (donde los items ya
   tienen sus IDs asignados después del import del FALLBACK con titles
   prefijados, ej. "BMW - Motor - Poleas").
2. Para cada item, calcula el title corto quitando el prefijo del parent:
   - Brand (sin parent): se queda como esta ("BMW", "Audi", ...).
   - Categoria (parent = brand): "BMW - Motor" -> "Motor".
   - Subcategoria (parent = categoria): "BMW - Motor - Poleas" -> "Poleas".
3. Genera un Excel de Matrixify con Command=MERGE y matching por
   Menu Item: ID (no por Title) para que no haya ambiguedad.

Uso:
  python3 scripts/15_generar_mega_menu_rename_corto.py [path/to/Export.xlsx]

Si no pasas argumento, busca el Export_*.xlsx mas reciente en menu/ que sea
distinto al backup viejo (Export_2026-05-13_173845.xlsx).

Output:
  menu/Embler-Mega-Menu-RENAME-corto.xlsx
"""

import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
MENU_DIR = ROOT / "menu"
OLD_BACKUP = "Export_2026-05-13_173845.xlsx"  # el backup pre-import; ignorar
OUT_PATH = MENU_DIR / "Embler-Mega-Menu-RENAME-corto.xlsx"


def find_latest_export(explicit=None):
    if explicit:
        p = Path(explicit)
        if not p.is_absolute():
            p = ROOT / p
        if not p.exists():
            raise FileNotFoundError(p)
        return p
    candidates = [
        p for p in MENU_DIR.glob("Export_*.xlsx")
        if p.name != OLD_BACKUP
    ]
    if not candidates:
        raise FileNotFoundError(
            f"No se encontro un Export_*.xlsx reciente en {MENU_DIR}. "
            f"Exporta el menu desde Matrixify y guarda el archivo en menu/."
        )
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0]


def load_export(path):
    """Devuelve filas del menu mega-menu como lista de dicts."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Menus"]
    headers = [c.value for c in ws[1]]
    col = {h: i for i, h in enumerate(headers)}
    required = [
        "Handle", "Title",
        "Menu Item: ID", "Menu Item: Title", "Menu Item: Parent ID",
    ]
    for r in required:
        if r not in col:
            raise ValueError(f"Falta la columna {r!r} en el export.")
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[col["Handle"]] != "mega-menu":
            continue
        rows.append({
            "id": row[col["Menu Item: ID"]],
            "title": row[col["Menu Item: Title"]],
            "parent_id": row[col["Menu Item: Parent ID"]],
            "menu_handle": row[col["Handle"]],
            "menu_title": row[col["Title"]],
        })
    return rows


def compute_short_titles(rows):
    """Para cada row, calcula el title corto quitando el prefix del parent."""
    by_id = {str(r["id"]): r for r in rows if r["id"] is not None}
    out = []
    for r in rows:
        rid = r["id"]
        title = r["title"] or ""
        parent_id = r["parent_id"]
        if not parent_id:
            short = title  # brand-level: no se toca
            level = "marca"
        else:
            parent = by_id.get(str(parent_id))
            parent_title = (parent["title"] if parent else "") or ""
            prefix = f"{parent_title} - "
            if title.startswith(prefix):
                short = title[len(prefix):]
            else:
                short = title  # fallback: dejar tal cual si no matchea
            level = "categoria" if (parent and not parent.get("parent_id_resolved")) else "subcategoria"
        out.append({
            "id": rid,
            "old_title": title,
            "new_title": short,
            "level": level,
            "menu_handle": r["menu_handle"],
            "menu_title": r["menu_title"],
            "parent_id": parent_id,
        })
    return out


def write_rename_xlsx(items):
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Menus"
    headers = [
        "Handle",
        "Command",
        "Title",
        "Menu Item: ID",
        "Menu Item: Command",
        "Menu Item: Title",
    ]
    ws.append(headers)
    changed = 0
    for it in items:
        if it["old_title"] == it["new_title"]:
            # No hay cambio (caso brand). Skip para no mandar updates inutiles.
            continue
        ws.append([
            it["menu_handle"],
            "MERGE",
            it["menu_title"],
            it["id"],
            "MERGE",
            it["new_title"],
        ])
        changed += 1
    wb.save(OUT_PATH)
    return changed


def write_preview(items):
    """Muestra ejemplos del rename para validar."""
    preview = MENU_DIR / "RENAME-PREVIEW.md"
    lines = []
    lines.append("# Preview del rename a labels cortos")
    lines.append("")
    lines.append("Esta es la transformacion que aplica `Embler-Mega-Menu-RENAME-corto.xlsx`:")
    lines.append("")
    lines.append("| ID | Antes | Después |")
    lines.append("|----|-------|---------|")
    for it in items:
        if it["old_title"] == it["new_title"]:
            continue
        lines.append(f"| `{it['id']}` | `{it['old_title']}` | `{it['new_title']}` |")
    preview.write_text("\n".join(lines), encoding="utf-8")
    return preview


def main():
    arg = sys.argv[1] if len(sys.argv) > 1 else None
    export_path = find_latest_export(arg)
    print(f"Usando export: {export_path.relative_to(ROOT)}")
    rows = load_export(export_path)
    print(f"Items del mega-menu: {len(rows)}")
    items = compute_short_titles(rows)
    changed = write_rename_xlsx(items)
    preview = write_preview(items)
    print(f"\nGenerado:")
    print(f"  {OUT_PATH.relative_to(ROOT)} ({changed} renames)")
    print(f"  {preview.relative_to(ROOT)} (preview de los cambios)")

    # Sanity check: contar por nivel
    sin_parent = sum(1 for it in items if not it["parent_id"])
    con_parent = sum(1 for it in items if it["parent_id"])
    print(f"\nDistribución: {sin_parent} marcas (sin cambio) + {con_parent} cat/subcat (renombradas)")


if __name__ == "__main__":
    main()
