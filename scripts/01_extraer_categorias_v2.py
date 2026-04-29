"""
Extrae los datos de input/CRUCE_ML_MC.xlsx (schema nuevo, 4 hojas) y los separa en CSVs por categoria.
Genera archivos en new-output/<hoja>/ listos para ser procesados con IA.

Schema nuevo (hojas ML_*):
  col 0: Id              col 12: URL Publicacion
  col 1: Situacion       col 13: Compatibilidades
  col 2: Categoria       col 14: Compatibilidades Restricciones
  col 3: Titulo          col 15: Marca
  col 4: Descripcion     col 16: Numero de parte
  col 5: Precio          col 17: Color manguera intercooler
  col 6: SKU             col 18: Tipo de vehiculo
  col 7: Estado          col 19: Origen
  col 8: Stock Tlalpan   col 20: Codigo OEM
  col 9: Disp. stock     col 21: Modelo
  col 10: Garantia       col 22: Lado
  col 11: Tags           col 23: CLAVES_COMBINADAS
  cols 24-27: match Microsip (MC_SKU, MC_ARTICULO_ID, MC_NOMBRE, MC_ESTATUS)
  cols 28-33: metadata match (n_claves, candidatos, TIENE_MATCH, AMBIGUO)

Hoja MC_sin_match: solo 11 cols del ERP (sin categoria), se vuelca como un solo CSV.
"""

import openpyxl
import csv
import os
from collections import defaultdict

INPUT_FILE = "input/CRUCE_ML_MC.xlsx"
OUTPUT_ROOT = "new-output"

# Indices del schema nuevo en hojas ML_*
COL_CATEGORIA = 2
COL_TITULO = 3
COL_MARCA = 15

MARCA_NORM = {
    "ORIGINAL FREY GERMAN TECHNOLOGY QUALITY": "Original Frey",
    "ORIGINAL FREY GERMAN TECHNOLOGY GERMAN Q": "Original Frey",
    "ORIGINAL FREY GERMAN TECHNOLOGY GERMAN QUALITY": "Original Frey",
    "ORIGINAL FREY GERMAN TECNHLOGY QUALITY": "Original Frey",
    "ORIGINAL FREY GERMAN TECHNOLOGY": "Original Frey",
    "EMBLER AUTOPARTES EUROPEAS": "Embler",
    "EMBLER": "Embler",
}

ML_SHEETS = ["ML_con_match", "ML_sin_match"]
MC_SHEET = None


def normalizar_marca(marca):
    if not marca:
        return ""
    marca_str = str(marca).strip()
    marca_upper = marca_str.upper()
    for key, val in MARCA_NORM.items():
        if key.upper() in marca_upper or marca_upper in key.upper():
            return val
    return marca_str.title() if marca_str.isupper() else marca_str


def limpiar_header(h):
    if not h:
        return ""
    return (
        str(h)
        .replace("_x000D_", " ")
        .replace("\r\n", " ")
        .replace("\n", " ")
        .replace("\x0d", " ")
        .strip()
    )


def limpiar_valor(val):
    if val is None:
        return ""
    return str(val).replace("\r\n", " ").replace("\n", " ").strip()


def extraer_subcategoria(categoria):
    if not categoria:
        return "otros"
    partes = str(categoria).split(">")
    if len(partes) >= 3:
        sub = partes[2].strip().lower()
    elif len(partes) >= 2:
        sub = partes[1].strip().lower()
    else:
        sub = partes[0].strip().lower()
    return sub


def clasificar_producto(categoria, titulo):
    cat = str(categoria).lower() if categoria else ""

    if "refacciones autos" in cat or "refacciones de auto" in cat:
        sub = extraer_subcategoria(categoria)
        if any(x in sub for x in ["motor", "admisi", "escape", "enfriamiento", "turbo"]):
            return "refacciones_motor"
        elif any(x in sub for x in ["suspensi", "direcci", "amortiguad"]):
            return "refacciones_suspension"
        elif any(x in sub for x in ["freno", "pastilla", "disco"]):
            return "refacciones_frenos"
        elif any(x in sub for x in ["transmisi", "clutch", "embrague"]):
            return "refacciones_transmision"
        elif any(x in sub for x in ["electri", "sensor", "encendido", "alternador", "bobina"]):
            return "refacciones_electrico"
        elif any(x in sub for x in ["carrocer", "espejo", "puerta", "defensa", "parrilla"]):
            return "refacciones_carroceria"
        elif any(x in sub for x in ["aire acondicionado", "calefacci", "clima"]):
            return "refacciones_clima"
        else:
            return "refacciones_otros"
    elif "accesorio" in cat and "tuning" not in cat:
        return "accesorios"
    elif "tuning" in cat:
        return "tuning"
    elif "moto" in cat:
        return "motos"
    elif "pesada" in cat or "linea pesada" in cat:
        return "linea_pesada"
    elif "herramienta" in cat:
        return "herramientas"
    else:
        return "otros"


def procesar_hoja_ml(wb, sheet_name, output_dir):
    ws = wb[sheet_name]

    headers_raw = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers_raw = list(row)
        break

    headers = [limpiar_header(h) for h in headers_raw]
    headers_extra = ["marca_normalizada", "subcategoria_limpia", "categoria_archivo"]
    all_headers = headers + headers_extra

    categorias = defaultdict(list)
    total = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        total += 1
        row_list = list(row)

        categoria_ml = row_list[COL_CATEGORIA] if len(row_list) > COL_CATEGORIA else ""
        titulo_ml = row_list[COL_TITULO] if len(row_list) > COL_TITULO else ""
        marca_ml = row_list[COL_MARCA] if len(row_list) > COL_MARCA else ""

        cat_archivo = clasificar_producto(categoria_ml, titulo_ml)
        marca_norm = normalizar_marca(marca_ml)
        subcat = extraer_subcategoria(categoria_ml)

        row_list.extend([marca_norm, subcat, cat_archivo])
        categorias[cat_archivo].append(row_list)

    os.makedirs(output_dir, exist_ok=True)

    print(f"\n[{sheet_name}] {total} filas leidas. Distribucion:")
    for cat_name, rows in sorted(categorias.items(), key=lambda x: -len(x[1])):
        output_path = os.path.join(output_dir, f"{cat_name}.csv")
        print(f"  {cat_name}: {len(rows):>6d} -> {output_path}")
        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(all_headers)
            for r in rows:
                writer.writerow([limpiar_valor(v) for v in r])

    return total


def procesar_hoja_mc(wb, sheet_name, output_dir):
    """MC_sin_match va completo a un solo CSV (no tiene categoria ML)."""
    ws = wb[sheet_name]

    headers_raw = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers_raw = list(row)
        break

    headers = [limpiar_header(h) for h in headers_raw]

    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, "mc_sin_match.csv")

    total = 0
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for row in ws.iter_rows(min_row=2, values_only=True):
            total += 1
            writer.writerow([limpiar_valor(v) for v in row])

    print(f"\n[{sheet_name}] {total} filas -> {output_path}")
    return total


def main():
    print(f"Leyendo {INPUT_FILE}...")
    wb = openpyxl.load_workbook(INPUT_FILE, read_only=True)

    os.makedirs(OUTPUT_ROOT, exist_ok=True)

    grand_total = 0
    for sheet in ML_SHEETS:
        if sheet not in wb.sheetnames:
            print(f"AVISO: hoja {sheet} no encontrada, saltando")
            continue
        out_dir = os.path.join(OUTPUT_ROOT, sheet.lower())
        grand_total += procesar_hoja_ml(wb, sheet, out_dir)

    if MC_SHEET and MC_SHEET in wb.sheetnames:
        out_dir = os.path.join(OUTPUT_ROOT, MC_SHEET.lower())
        grand_total += procesar_hoja_mc(wb, MC_SHEET, out_dir)

    wb.close()

    print(f"\n{'='*60}")
    print(f"Total filas procesadas: {grand_total}")
    print(f"Output: {OUTPUT_ROOT}/")
    for sub in sorted(os.listdir(OUTPUT_ROOT)):
        sub_path = os.path.join(OUTPUT_ROOT, sub)
        if os.path.isdir(sub_path):
            files = [f for f in sorted(os.listdir(sub_path)) if f.endswith(".csv")]
            print(f"  {sub}/ ({len(files)} archivos)")


if __name__ == "__main__":
    main()
