"""
Consolida los 3 Import_Result de Matrixify en un solo Excel de errores
y reorganiza todo bajo outputs/collections-matrixify/.

Estructura final:
  outputs/collections-matrixify/
    source/Embler-Collections.xlsx          (input original a Matrixify)
    bloques/Embler-Collections-Bloque-2.xlsx
    bloques/Embler-Collections-Bloque-3.xlsx
    import-results/Import_Result_2026-05-13_074131.xlsx  (bloque 1)
    import-results/Import_Result_2026-05-13_132357.xlsx  (bloque 2)
    import-results/Import_Result_2026-05-13_133148.xlsx  (bloque 3)
    errores/Embler-Collections-Errores.xlsx              (16 handles fallidos)

El archivo de errores incluye:
- Todas las filas (con reglas) de los handles que fallaron.
- Columnas extra: Bloque, Matrixify Error.
"""

import shutil
from pathlib import Path
from collections import defaultdict

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SOURCE = ROOT / "outputs/2026-05-12-collections-matrixify/Embler-Collections.xlsx"
RESULTS = [
    ("Bloque 1", ROOT / "Import_Result_2026-05-13_074131.xlsx"),
    ("Bloque 2", ROOT / "Import_Result_2026-05-13_132357.xlsx"),
    ("Bloque 3", ROOT / "Import_Result_2026-05-13_133148.xlsx"),
]
DEST = ROOT / "outputs/collections-matrixify"
SHEET = "Smart Collections"


def load_source():
    wb = openpyxl.load_workbook(SOURCE, data_only=False)
    ws = wb[SHEET]
    headers = [c.value for c in ws[1]]
    rows_by_handle = defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        h = row[0]
        if h:
            rows_by_handle[h].append(row)
    return headers, rows_by_handle


def collect_failed():
    """Devuelve [(bloque, handle, comment)] preservando el orden y eliminando duplicados por handle."""
    seen = set()
    out = []
    for label, path in RESULTS:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[SHEET]
        # Una entrada por handle: tomamos el primer comentario de error.
        first_error_per_handle = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            h, result, comment = row[0], row[16], row[17]
            if not h or (isinstance(h, str) and h.startswith("###")):
                continue
            if result == "Failed" and h not in first_error_per_handle:
                first_error_per_handle[h] = comment
        for h, comment in first_error_per_handle.items():
            if h in seen:
                continue
            seen.add(h)
            out.append((label, h, comment))
    return out


def write_errores(headers, rows_by_handle, failed):
    out_path = DEST / "errores/Embler-Collections-Errores.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.append(["Bloque", "Matrixify Error"] + headers)
    for label, h, comment in failed:
        for src_row in rows_by_handle[h]:
            ws.append([label, comment] + list(src_row))
    wb.save(out_path)
    return out_path


def reorganize():
    """Copia archivos a la nueva estructura. No borra los originales aún."""
    moves = [
        (SOURCE, DEST / "source/Embler-Collections.xlsx"),
        (ROOT / "outputs/2026-05-13-collections-bloques/Embler-Collections-Bloque-2.xlsx",
         DEST / "bloques/Embler-Collections-Bloque-2.xlsx"),
        (ROOT / "outputs/2026-05-13-collections-bloques/Embler-Collections-Bloque-3.xlsx",
         DEST / "bloques/Embler-Collections-Bloque-3.xlsx"),
    ]
    for label, path in RESULTS:
        moves.append((path, DEST / "import-results" / path.name))

    for src, dst in moves:
        dst.parent.mkdir(parents=True, exist_ok=True)
        if src.exists():
            shutil.copy2(src, dst)
            print(f"  Copiado: {src.relative_to(ROOT)} -> {dst.relative_to(ROOT)}")
        else:
            print(f"  [!] No existe: {src}")


def main():
    headers, rows_by_handle = load_source()
    failed = collect_failed()
    print(f"Handles fallidos en total: {len(failed)}")
    for label, h, comment in failed:
        print(f"  [{label}] {h}: {comment[:80]}...")

    err_path = write_errores(headers, rows_by_handle, failed)
    print(f"\nErrores -> {err_path.relative_to(ROOT)}")

    print("\nReorganizando archivos:")
    reorganize()


if __name__ == "__main__":
    main()
