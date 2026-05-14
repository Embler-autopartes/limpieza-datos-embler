"""
Genera el archivo Matrixify para reemplazar el "Mega menu" de Shopify
con la estructura: marca -> categoria (group) -> max 3 subcategorias
(top 3 alfabeticas por group).

Reglas:
- Marca (nivel 1): SIN link. Usamos Resource Type=HTTP + URL=# (placeholder).
- Categoria (nivel 2): linkea a la collection brand-group.
- Subcategoria (nivel 3): linkea a la collection brand-group-subgroup.
- Titulos prefijados con la marca para que el parent matching sea univoco
  (ej. "BMW - Motor", "BMW - Motor - Poleas").

Branding final del menu (13 marcas):
- KEEP labels actuales: BMW, Mercedes Benz, Audi, Mini Cooper, Porsche, Smart, Volkswagen
- ADD nuevas (orden alfabetico al final): Bentley, Fiat, Jaguar, Land Rover, Seat, Volvo
- BORRAR del menu (no aparecen en el output): Sprinter, Citroen, Peugeot, Renault

Output:
- menu/Embler-Mega-Menu.xlsx (Matrixify import)
- menu/MEGA-MENU-PROPUESTA.md (preview de la estructura para validar antes de importar)
- menu/README.md (instrucciones de import)
"""

from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SOURCE = ROOT / "Procesamiento de Catálogo/outputs/collections-matrixify/source/Embler-Collections.xlsx"
OUT_DIR = ROOT / "menu"
OUT_XLSX = OUT_DIR / "Embler-Mega-Menu.xlsx"
OUT_XLSX_FALLBACK = OUT_DIR / "Embler-Mega-Menu-FALLBACK-con-prefijos.xlsx"
OUT_PREVIEW = OUT_DIR / "MEGA-MENU-PROPUESTA.md"
OUT_README = OUT_DIR / "README.md"

MENU_TITLE = "Mega menu"
MENU_HANDLE = "mega-menu"
SUB_PER_GROUP = 3  # tope de subcategorias por categoria

# Display label en el menu -> nombre del metafield brand (collection prefix)
# Si difieren, el menu mantiene el display label pero linkea a la collection real.
BRAND_LABELS = [
    # Existentes (orden actual del menu)
    ("BMW", "BMW", "bmw"),
    ("Mercedes Benz", "Mercedes-Benz", "mercedes-benz"),
    ("Audi", "Audi", "audi"),
    ("Mini Cooper", "Mini", "mini"),
    ("Porsche", "Porsche", "porsche"),
    ("Smart", "Smart", "smart"),
    ("Volkswagen", "Volkswagen", "volkswagen"),
    # Nuevas (alfabeticas)
    ("Bentley", "Bentley", "bentley"),
    ("Fiat", "Fiat", "fiat"),
    ("Jaguar", "Jaguar", "jaguar"),
    ("Land Rover", "Land Rover", "land-rover"),
    ("Seat", "Seat", "seat"),
    ("Volvo", "Volvo", "volvo"),
]


def parse_collections():
    """handle -> {brand, group, sub_group, title}"""
    wb = openpyxl.load_workbook(SOURCE, data_only=True)
    ws = wb["Smart Collections"]
    collections = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        handle, _cmd, title = row[0], row[1], row[2]
        rule_col, _, rule_cond = row[11], row[12], row[13]
        if not handle:
            continue
        c = collections.setdefault(handle, {"title": title, "brand": None, "group": None, "sub_group": None})
        if rule_col == "Metafield: global._brand":
            c["brand"] = rule_cond
        elif rule_col == "Metafield: global.group":
            c["group"] = rule_cond
        elif rule_col == "Metafield: global.sub_group":
            c["sub_group"] = rule_cond
    return collections


def build_hierarchy(collections):
    """Devuelve {brand_metafield_value: {group: {'handle': ..., 'subs': [(sg, handle), ...]}}}"""
    brand_group_handle = defaultdict(dict)
    brand_group_subs = defaultdict(lambda: defaultdict(list))
    for h, c in collections.items():
        b, g, sg = c["brand"], c["group"], c["sub_group"]
        if not b:
            continue
        if g and not sg:
            brand_group_handle[b][g] = h
        elif g and sg:
            brand_group_subs[b][g].append((sg, h))
    # Order subs alphabetically and cap to SUB_PER_GROUP
    out = {}
    for b, groups in brand_group_handle.items():
        out[b] = {}
        for g, h in sorted(groups.items()):
            subs_sorted = sorted(brand_group_subs[b][g], key=lambda x: x[0].lower())
            out[b][g] = {"handle": h, "subs": subs_sorted[:SUB_PER_GROUP], "subs_total": len(brand_group_subs[b][g])}
    return out


def make_row(item_title, resource_type, resource_handle, url, parent_title, position):
    """Construye una fila con todas las columnas del schema Matrixify."""
    return {
        "Handle": MENU_HANDLE,
        "Command": "REPLACE",
        "Title": MENU_TITLE,
        "Menu Item: Title": item_title,
        "Menu Item: Command": "MERGE",
        "Menu Item: Resource Type": resource_type,
        "Menu Item: Resource Handle": resource_handle,
        "Menu Item: URL": url,
        "Menu Item: Parent Title": parent_title,
        "Menu Item: Position": position,
    }


def build_rows(hierarchy, use_prefix=False):
    """Genera las filas para el sheet Menus.

    Command=REPLACE en todas las filas: Matrixify borra el menu actual
    (handle `mega-menu`) y lo recrea solo con los items de este archivo.
    Esto elimina automaticamente Sprinter, Citroen, Peugeot, Renault, y
    los hijos sueltos actuales de BMW sin limpieza manual.

    Resource Type=COLLECTION + Resource Handle hace que Shopify genere
    el link /collections/<handle> automaticamente (sin necesidad de
    llenar URL).

    use_prefix=True genera titulos con prefijo de marca para que sean
    globalmente unicos ("BMW - Motor"). use_prefix=False usa titulos
    cortos ("Motor") y apuesta a que Matrixify resuelve el parent
    matching por orden de proceso (filas agrupadas por marca).
    """
    rows = []
    pos_brand = 0
    for label, brand_key, brand_handle in BRAND_LABELS:
        pos_brand += 1
        # Nivel 1: marca (sin link real - HTTP + #)
        rows.append(make_row(
            item_title=label,
            resource_type="HTTP",
            resource_handle="",
            url="#",
            parent_title="",
            position=pos_brand,
        ))
        # Nivel 2: categorias del brand
        groups = hierarchy.get(brand_key, {})
        for pos_g, (group, ginfo) in enumerate(sorted(groups.items()), start=1):
            cat_title = f"{label} - {group}" if use_prefix else group
            rows.append(make_row(
                item_title=cat_title,
                resource_type="COLLECTION",
                resource_handle=ginfo["handle"],
                url="",
                parent_title=label,
                position=pos_g,
            ))
            # Nivel 3: subcategorias (tope 3 alfabeticas)
            for pos_sg, (sg, sg_handle) in enumerate(ginfo["subs"], start=1):
                sub_title = f"{label} - {group} - {sg}" if use_prefix else sg
                rows.append(make_row(
                    item_title=sub_title,
                    resource_type="COLLECTION",
                    resource_handle=sg_handle,
                    url="",
                    parent_title=cat_title,
                    position=pos_sg,
                ))
    return rows


def write_xlsx(rows, out_path):
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Menus"
    headers = [
        "Handle",
        "Command",
        "Title",
        "Menu Item: Title",
        "Menu Item: Command",
        "Menu Item: Resource Type",
        "Menu Item: Resource Handle",
        "Menu Item: URL",
        "Menu Item: Parent Title",
        "Menu Item: Position",
    ]
    ws.append(headers)
    for r in rows:
        ws.append([r[h] for h in headers])
    wb.save(out_path)


def write_preview(hierarchy):
    out = []
    out.append("# Mega menu — estructura propuesta para Shopify")
    out.append("")
    out.append("Estructura que se importará a Matrixify (archivo `Embler-Mega-Menu.xlsx`). Valida antes de subir.")
    out.append("")
    out.append("**Reglas:**")
    out.append("- Marca (nivel 1): SIN link (placeholder `#`).")
    out.append("- Categoria (nivel 2): link a la collection `brand-group`.")
    out.append("- Subcategoria (nivel 3): link a la collection `brand-group-subgroup`, máximo 3 por categoria (orden alfabetico).")
    out.append("")
    out.append("**Marcas borradas del menu actual:** Sprinter, Citroën, Peugeot, Renault (no tienen collections).")
    out.append("")
    out.append("**Marcas agregadas:** Bentley, Fiat, Jaguar, Land Rover, Seat, Volvo.")
    out.append("")
    out.append("---")
    out.append("")
    total_items = 0
    total_brands = 0
    total_cats = 0
    total_subs = 0
    for label, brand_key, brand_handle in BRAND_LABELS:
        groups = hierarchy.get(brand_key, {})
        total_brands += 1
        n_cat = len(groups)
        n_sub_capped = sum(len(g["subs"]) for g in groups.values())
        n_sub_total = sum(g["subs_total"] for g in groups.values())
        total_cats += n_cat
        total_subs += n_sub_capped
        total_items += 1 + n_cat + n_sub_capped
        out.append(f"## {label}")
        out.append(f"**Sin link** · {n_cat} categorias · {n_sub_capped} subcategorias en menu (de {n_sub_total} totales)")
        out.append("")
        if not groups:
            out.append("_Sin contenido._")
            out.append("")
            continue
        for group, ginfo in sorted(groups.items()):
            out.append(f"- **{label} - {group}** → `/collections/{ginfo['handle']}`")
            for sg, sg_handle in ginfo["subs"]:
                out.append(f"  - {sg} → `/collections/{sg_handle}`")
            if ginfo["subs_total"] > SUB_PER_GROUP:
                out.append(f"  - _… +{ginfo['subs_total'] - SUB_PER_GROUP} subcategorias no incluidas (cap de {SUB_PER_GROUP})_")
        out.append("")
    out.insert(11, "")
    out.insert(11, f"**Total items en el menu:** {total_items} ({total_brands} marcas + {total_cats} categorias + {total_subs} subcategorias)")
    OUT_PREVIEW.write_text("\n".join(out), encoding="utf-8")
    return total_items


def write_readme(total_items):
    txt = f"""# Mega menu — import a Shopify via Matrixify

## Archivos

- `Embler-Mega-Menu.xlsx` — **archivo principal a importar**. Labels cortos (`Motor`, `Poleas`, etc.).
- `Embler-Mega-Menu-FALLBACK-con-prefijos.xlsx` — fallback con labels prefijados (`BMW - Motor`, `BMW - Motor - Poleas`). Usa este SI el principal rompe la jerarquia parent-child (ver caveat de "Parent matching" abajo).
- `MEGA-MENU-PROPUESTA.md` — preview de la estructura para validar antes del import.
- `Export_2026-05-13_173845.xlsx` — backup del menu actual (para revertir si algo sale mal).
- `ESTRUCTURA-MENU.md` — referencia completa de las 895 collections (no solo las del menu).

## Resumen

- Total items en el menu: **{total_items}**
- Niveles: marca (sin link) → categoria → max 3 subcategorias
- Menu target: `mega-menu` (Title: "Mega menu")

## Cómo importar

> **Importante:** el archivo usa `Command=REPLACE` a nivel del menu. Eso significa que Matrixify
> **borra completamente el menu "Mega menu" actual** y lo recrea solo con los items de este archivo.
> Resultado: Sprinter, Citroën, Peugeot, Renault, y cualquier hijo viejo de BMW desaparecen
> automaticamente. No es un MERGE.

1. **Backup recomendado:** en Shopify Admin → Apps → Matrixify → Export, exporta el menu actual ("Mega menu") por si quieres revertir.
2. Matrixify → Import → sube `Embler-Mega-Menu.xlsx`.
3. Run import. El menu se elimina y se recrea desde cero con los 299 items.
4. Verifica en `/admin/content/menus/` que el nuevo "Mega menu" aparezca con la estructura esperada.

**Nota sobre el ID del menu:** REPLACE borra el menu y crea uno nuevo, asi que el ID interno
(`313377915250` en la URL actual) puede cambiar. El handle (`mega-menu`) **se mantiene**, asi que
si tu theme referencia el menu por handle (lo normal), todo sigue funcionando. Si el theme
referencia por ID, hay que actualizarlo.

## Caveats

- **Parent matching por Title (riesgo manejable)**: Matrixify identifica el parent por Title exacto.
  El archivo principal usa labels cortos (`Motor`, `Suspensión`), lo que significa que **multiples
  marcas tienen una categoria llamada "Motor"**. Para que funcione, las filas estan agrupadas
  hierarquicamente por marca (BMW completo, luego Audi completo, ...) apostando a que Matrixify
  procesa en orden y asigna parent al match mas reciente.

  **Si después del import notas que subcategorias quedaron bajo la marca equivocada** (ej. ves
  "Poleas" bajo BMW cuando deberia estar bajo Audi), no funciono el matching posicional. En ese caso:
  1. Importa `Export_2026-05-13_173845.xlsx` para revertir al menu de antes.
  2. Importa `Embler-Mega-Menu-FALLBACK-con-prefijos.xlsx` que usa titulos globalmente unicos.
  3. Los labels en el menu apareceran con prefijo de marca (verbose), pero estructura correcta.

- **Marca sin link**: usamos Resource Type=HTTP + URL=`#`. Algunos themes tratan esto como hover-only (correcto para mega menus); otros pueden marcarlo como link roto. Si el theme se queja, cambia el `#` por la collection padre (`bmw`, `audi`, etc.) en el admin.
- **Subcategorias topadas a 3 alfabeticas**: muchas categorias tienen >3 (ej. BMW Motor tiene 30+ subs). El cap es para no rebasar el limite practico de Shopify. Si quieres priorizar otras subcategorias (las mas vendidas, las top), editalas en el admin o regenera el archivo cambiando `SUB_PER_GROUP` o la regla de orden en `scripts/14_generar_mega_menu_matrixify.py`.

## Para regenerar

```bash
python3 scripts/14_generar_mega_menu_matrixify.py
```
"""
    OUT_README.write_text(txt, encoding="utf-8")


def main():
    collections = parse_collections()
    hierarchy = build_hierarchy(collections)

    rows_short = build_rows(hierarchy, use_prefix=False)
    write_xlsx(rows_short, OUT_XLSX)

    rows_prefix = build_rows(hierarchy, use_prefix=True)
    write_xlsx(rows_prefix, OUT_XLSX_FALLBACK)

    total_items = write_preview(hierarchy)
    write_readme(total_items)
    print(f"Generado:")
    print(f"  {OUT_XLSX.relative_to(ROOT)} ({len(rows_short)} filas) - labels CORTOS (recomendado)")
    print(f"  {OUT_XLSX_FALLBACK.relative_to(ROOT)} ({len(rows_prefix)} filas) - labels con PREFIJO (fallback)")
    print(f"  {OUT_PREVIEW.relative_to(ROOT)}")
    print(f"  {OUT_README.relative_to(ROOT)}")
    print(f"Total items en el menu: {total_items}")


if __name__ == "__main__":
    main()
