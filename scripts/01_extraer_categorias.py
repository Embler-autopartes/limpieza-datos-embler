"""
Extrae los datos del archivo CRUCE y los separa en CSVs por categoria.
Genera archivos en output/ listos para ser procesados con IA.
"""

import openpyxl
import csv
import os
import re
from collections import defaultdict

INPUT_FILE = "input/INPUT.xlsx"
OUTPUT_DIR = "output"

# Mapeo de normalizacion de marcas
MARCA_NORM = {
    "ORIGINAL FREY GERMAN TECHNOLOGY QUALITY": "Original Frey",
    "ORIGINAL FREY GERMAN TECHNOLOGY GERMAN Q": "Original Frey",
    "ORIGINAL FREY GERMAN TECHNOLOGY GERMAN QUALITY": "Original Frey",
    "ORIGINAL FREY GERMAN TECNHLOGY QUALITY": "Original Frey",
    "ORIGINAL FREY GERMAN TECHNOLOGY": "Original Frey",
    "EMBLER AUTOPARTES EUROPEAS": "Embler",
    "EMBLER": "Embler",
}


def normalizar_marca(marca):
    if not marca:
        return ""
    marca_str = str(marca).strip()
    marca_upper = marca_str.upper()
    for key, val in MARCA_NORM.items():
        if key.upper() in marca_upper or marca_upper in key.upper():
            return val
    # Capitalizar si no esta en el mapeo
    return marca_str.title() if marca_str.isupper() else marca_str


def limpiar_header(h):
    """Limpia nombres de columna con caracteres especiales."""
    if not h:
        return ""
    return str(h).replace("\r\n", " ").replace("\n", " ").replace("\x0d", " ").strip()


def extraer_subcategoria(categoria):
    """Extrae la subcategoria del path de categoria de ML."""
    if not categoria:
        return "otros"
    partes = str(categoria).split(">")
    if len(partes) >= 3:
        sub = partes[2].strip().lower()
    elif len(partes) >= 2:
        sub = partes[1].strip().lower()
    else:
        sub = partes[0].strip().lower()
    return sub


def clasificar_producto(categoria, titulo):
    """Clasifica un producto en una de las categorias principales."""
    cat = str(categoria).lower() if categoria else ""
    tit = str(titulo).lower() if titulo else ""

    if "refacciones autos" in cat or "refacciones de auto" in cat:
        # Subdividir refacciones
        sub = extraer_subcategoria(categoria)
        if any(x in sub for x in ["motor", "admisi", "escape", "enfriamiento", "turbo"]):
            return "refacciones_motor"
        elif any(x in sub for x in ["suspensi", "direcci", "amortiguad"]):
            return "refacciones_suspension"
        elif any(x in sub for x in ["freno", "pastilla", "disco"]):
            return "refacciones_frenos"
        elif any(x in sub for x in ["transmisi", "clutch", "embrague"]):
            return "refacciones_transmision"
        elif any(x in sub for x in ["electri", "sensor", "encendido", "alternador", "bobina"]):
            return "refacciones_electrico"
        elif any(x in sub for x in ["carrocer", "espejo", "puerta", "defensa", "parrilla"]):
            return "refacciones_carroceria"
        elif any(x in sub for x in ["aire acondicionado", "calefacci", "clima"]):
            return "refacciones_clima"
        else:
            return "refacciones_otros"
    elif "accesorio" in cat and "tuning" not in cat:
        return "accesorios"
    elif "tuning" in cat:
        return "tuning"
    elif "moto" in cat:
        return "motos"
    elif "pesada" in cat or "línea pesada" in cat:
        return "linea_pesada"
    elif "herramienta" in cat:
        return "herramientas"
    else:
        return "otros"


def main():
    print(f"Leyendo {INPUT_FILE}...")
    wb = openpyxl.load_workbook(INPUT_FILE, read_only=True)
    ws = wb["Sheet1"]

    # Leer headers
    headers_raw = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers_raw = list(row)

    headers = [limpiar_header(h) for h in headers_raw]
    # Agregar columnas nuevas
    headers_extra = ["marca_normalizada", "subcategoria_limpia", "categoria_archivo"]
    all_headers = headers + headers_extra

    # Agrupar filas por categoria
    categorias = defaultdict(list)
    total = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        total += 1
        row_list = list(row)

        # Columnas: 13=Categoria_ML, 14=Titulo_ML, 26=Marca_ML
        categoria_ml = row_list[13] if len(row_list) > 13 else ""
        titulo_ml = row_list[14] if len(row_list) > 14 else ""
        marca_ml = row_list[26] if len(row_list) > 26 else ""

        cat_archivo = clasificar_producto(categoria_ml, titulo_ml)
        marca_norm = normalizar_marca(marca_ml)
        subcat = extraer_subcategoria(categoria_ml)

        row_list.extend([marca_norm, subcat, cat_archivo])
        categorias[cat_archivo].append(row_list)

    wb.close()

    print(f"Total filas leidas: {total}")
    print(f"\nDistribucion por categoria:")

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    for cat_name, rows in sorted(categorias.items(), key=lambda x: -len(x[1])):
        output_path = os.path.join(OUTPUT_DIR, f"{cat_name}.csv")
        print(f"  {cat_name}: {len(rows):>6d} filas -> {output_path}")

        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(all_headers)
            for row in rows:
                # Limpiar valores None y caracteres especiales
                cleaned = []
                for val in row:
                    if val is None:
                        cleaned.append("")
                    else:
                        cleaned.append(
                            str(val).replace("\r\n", " ").replace("\n", " ").strip()
                        )
                writer.writerow(cleaned)

    # Generar resumen
    print(f"\nArchivos generados en {OUTPUT_DIR}/:")
    for fname in sorted(os.listdir(OUTPUT_DIR)):
        if fname.endswith(".csv"):
            fpath = os.path.join(OUTPUT_DIR, fname)
            size = os.path.getsize(fpath)
            print(f"  {fname}: {size / 1024:.0f} KB")

    print("\nListo. Los archivos CSV estan listos para procesamiento con IA.")


if __name__ == "__main__":
    main()
